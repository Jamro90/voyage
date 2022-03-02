from openpyxl import *
from tkinter import filedialog, messagebox, Label, Frame
import os
from openpyxl import Workbook, load_workbook

def import_data():

    # inspecting data
    try:
        # data finding and reading
        data = filedialog.askopenfilename(filetypes = (("excel workbook", "*.xlsx"), ("json text","*.json"), ("all files", "*")))
        read_file = data.split(".")
        
        # reading xlsx
        if read_file[1] == "xlsx":
            wb = load_workbook(data)
            ws = wb.active
            pchor_list = []
        
            # Data importing
            for row in range(1, 200):
                if str(ws[chr(65) + str(row)].value) != "None":
                    pchor_list.append(f"{ws[chr(65) + str(row)].value},{ws[chr(67) + str(row)].value},{ws[chr(66) + str(row)].value}")
                else:
                    break
            wb.save(data)
            return pchor_list 

        # reading json
        elif read_file[1] == "json":
            print("json")
        
        # returning errors
        else:
            messagebox.showerror("ERROR!", "Nieprawidłowe rozszerzenie danych!")
    except:
        messagebox.showerror("ERROR!", "Nie znaleziono danych!")


def save_as(data):
    pass

def check_all():
    pass

def uncheck_all():
    pass

